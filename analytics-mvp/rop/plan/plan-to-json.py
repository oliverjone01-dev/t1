#!/usr/bin/env python3
# Конвертер плана РОП: Excel (формат ROP_plan_template) -> plan.json для build-rop.
# Запуск: python3 rop/plan/plan-to-json.py [путь.xlsx]
# По умолчанию берёт ROP_plan_template.xlsx рядом. Источник истины - Google Таблица,
# при обновлении: скачать как .xlsx, положить сюда, прогнать скрипт, пересобрать build-rop.
import sys, json, re, os, datetime
from openpyxl import load_workbook

here = os.path.dirname(os.path.abspath(__file__))
src = sys.argv[1] if len(sys.argv) > 1 else os.path.join(here, "ROP_plan_template.xlsx")
ws = load_workbook(src, data_only=True).active

def num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[^\d.,-]", "", str(v)).replace(" ", "")
    if not s: return None
    s = s.replace(",", ".") if s.count(",") == 1 and "." not in s else s.replace(",", "")
    try: return float(s)
    except: return None

def pct(v):
    n = num(v)
    if n is None: return None
    return n/100 if n > 1.5 else n   # 16 -> 0.16; 0.16 -> 0.16

months = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if not (isinstance(a, str) and re.fullmatch(r"\d{4}-\d{2}", a.strip())): continue
    m = a.strip()
    rev, chk, cr2, cr1 = num(ws.cell(r,2).value), num(ws.cell(r,3).value), pct(ws.cell(r,4).value), pct(ws.cell(r,5).value)
    leads, deals = num(ws.cell(r,6).value), num(ws.cell(r,7).value)
    if rev is None: continue
    months.append({"month": m, "rev": round(rev), "check": round(chk) if chk else None,
                   "cr2": cr2, "cr1": cr1,
                   "leads": round(leads) if leads else None, "deals": round(deals) if deals else None})

out = {"source": os.path.basename(src), "generatedAt": datetime.datetime.utcnow().isoformat()+"Z", "months": months}
dst = os.path.join(here, "plan.json")
json.dump(out, open(dst, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"plan.json: {len(months)} месяцев -> {dst}")
for x in months[:6]: print(" ", x)
